import json
import sys
import openpyxl
import os
import re
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

#--------#
# CONFIG #
#--------#

# Required
EXCEL_FILEPATH = 'items.xlsx'
SCRIPT_DIRECTORY_ABS_PATH = r'D:\Users\path\to\Price Tracker'

# Booleans
HIDE_CHROME_UI = False  # Set to True if you want to hide the Chrome UI while it webscrapes (can lead to issues with some sites)
PRINT_TO_TXT = True  # Set to True if you want the print() statements to print to a txt file instead of the console (useful if you run this on an auto-scheduler)

# Notion integration
NOTION_SECRET = 'secret_xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx'
NOTION_DATABASE_ID = 'xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx'

#------#
# CODE #
#------#

def alert(items_with_price_drop: list[dict]) -> None:
    """Alerts the user about items with price drops.

    This function posts a message to a Notion page using the Notion API.

    If you would like to be alerted a different way (e.g. print to console, email), then modify this function.

    Args:
        items_with_price_drop: List of items have that dropped in price
    """
    message = "Price dropped for the following items:\n"

    for item in items_with_price_drop:
        message += f"{item['name']} is now ${item['current_price']:.2f} (was ${item['last_price']:.2f}). Link: {item['url']}\n"

    endpoint = "https://api.notion.com/v1/pages"

    headers = {
        "Authorization": f"Bearer {NOTION_SECRET}",
        "Content-Type": "application/json",
        "Notion-Version": "2022-06-28"
    }

    body = {
        "parent": {
            "database_id": NOTION_DATABASE_ID
        },
        "properties": {
            "Name": {
                "title": [
                    {
                        "text": {
                            "content": "Price tracker"
                        }
                    }
                ]
            },
            "Status": {
                "select": {
                    "name": "High priority"
                }
            }
        },
        "children": [
            {
                "object": "block",
                "paragraph": {
                    "rich_text": [
                        {
                            "text": {
                                "content": message
                            }
                        }
                    ]
                }
            }
        ]
    }

    response = requests.post(endpoint, headers=headers, data=json.dumps(body))

    if response.status_code == 200:
        print("Message successfully posted to Notion!")

    else:
        print(f"Failed to post message to Notion. Status code: {response.status_code}, Response: {response.text}")


def get_price(driver, item) -> float:
    """Gets the price of the item using the web driver.
    
    Args:
        driver: The web driver
        item: The item, which must contain a list of comma-delimited CSS selectors
        
    Returns:
        float: The items' price
    """
    
    driver.get(item['url'])

    # Explicitly wait until the element is present
    price_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, item['css_selector']))
    )

    # Use a regex pattern to match any numeric value including decimal points
    matches = re.findall(r'\d+\.\d+|\d+', price_element.text)

    if matches:
        return min(float(price) for price in matches)
    
    else:
        raise ValueError(f"no price found for {item['name']} at {item['url']}")


def get_driver():

    # Options to suppress logging
    chrome_options = Options()
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    
    if HIDE_CHROME_UI:
        chrome_options.add_argument("--headless")

    return webdriver.Chrome(options=chrome_options)


def main():

    os.chdir(SCRIPT_DIRECTORY_ABS_PATH)
    
    with open('output.txt', 'w') as f:

        if PRINT_TO_TXT:
            sys.stdout = f

        driver = get_driver()

        workbook = openpyxl.load_workbook(EXCEL_FILEPATH)
        sheet = workbook.active

        items_with_price_drop = []

        # Assuming the data starts on row 2 (row 1 being the header)
        for row in range(2, sheet.max_row + 1):

            item = {
                'name': sheet.cell(row=row, column=1).value,
                'url': sheet.cell(row=row, column=2).value,
                'css_selector': sheet.cell(row=row, column=3).value,
                'last_price': sheet.cell(row=row, column=4).value
            }

            try:

                current_price = get_price(driver, item)
                sheet.cell(row=row, column=4).value = current_price

                print(f"{item['name']} costs ${current_price:.2f}")

                if item['last_price'] is None:
                    item['last_price'] = current_price

                if current_price < float(item['last_price']):
                    item['current_price'] = current_price
                    items_with_price_drop.append(item)

            except Exception as e:
                print(f"Failed to get price for {item['name']}: {e}")
                continue

        workbook.save(EXCEL_FILEPATH)
        driver.quit()

        if items_with_price_drop:
            alert(items_with_price_drop)


if __name__ == "__main__":
    main()
